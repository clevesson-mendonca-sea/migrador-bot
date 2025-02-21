import os
import discord
from dotenv import load_dotenv
from discord.ext import commands, tasks
import requests
from bs4 import BeautifulSoup
import pandas as pd
import logging
import aiohttp
import asyncio
import re
from urllib.parse import urljoin, urlparse
import time
from collections import defaultdict
import io

# Configuração de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Carregar variáveis de ambiente
load_dotenv()
DISCORD_TOKEN = os.getenv('DISCORD_TOKEN')
MONDAY_API_KEY = os.getenv('MONDAY_API_KEY')
BOARD_ID = 2422124618
DISCORD_CHANNEL_ID = 1305193863731871764 

# Configuração do bot
intents = discord.Intents.default()
intents.messages = True
intents.message_content = True
bot = commands.Bot(command_prefix='!', intents=intents)

# Fila de comandos
command_queue = asyncio.Queue()

processed_items = set()

# Controlar taxas de comandos
last_command_time = defaultdict(float)
command_cooldown = 5

# Inicializando a sessão aiohttp
client_session = None

async def get_board_updates():
    url = "https://api.monday.com/v2"
    headers = {
        "Authorization": MONDAY_API_KEY,
        "Content-Type": "application/json"
    }
    query = """
    {
        boards(ids: %d) {
            items {
                id
                name
                created_at
            }
        }
    }
    """ % BOARD_ID
    
    async with client_session.post(url, headers=headers, json={'query': query}) as response:
        response.raise_for_status()
        data = await response.json()
        return data

async def notify_new_task(task_name):
    channel = bot.get_channel(DISCORD_CHANNEL_ID)
    if channel:
        await channel.send(f"Novo chamado criado: {task_name}")

@tasks.loop(seconds=60)  # Verifica atualizações a cada 60 segundos
async def check_for_updates():
    try:
        updates = await get_board_updates()
        items = updates.get('data', {}).get('boards', [])[0].get('items', [])
        
        for item in items:
            task_id = item['id']
            task_name = item['name']
            created_at = item['created_at']

            # Verifica se o item já foi processado
            if task_id not in processed_items:
                processed_items.add(task_id)
                logging.info(f"Novo item encontrado: {task_name} (Criado em: {created_at})")
                await notify_new_task(task_name)

    except Exception as e:
        logging.error(f"Erro ao verificar atualizações do board: {e}")

async def command_processor():
    while True:
        ctx, command = await command_queue.get()
        try:
            await command(ctx)
        finally:
            command_queue.task_done()

@bot.event
async def on_ready():
    global client_session
    client_session = aiohttp.ClientSession()
    bot.loop.create_task(command_processor())
    logging.info(f'Bot {bot.user.name} está conectado e pronto para uso!')

def rate_limited(func):
    async def wrapper(ctx, *args, **kwargs):
        now = time.time()
        if now - last_command_time[ctx.author.id] < command_cooldown:
            await ctx.send("Por favor, aguarde antes de enviar outro comando.")
            return
        last_command_time[ctx.author.id] = now
        return await func(ctx, *args, **kwargs)
    return wrapper

async def get_page(url):
    async with client_session.get(url, timeout=30, ssl=False) as response:
        response.raise_for_status()
        
        content = await response.read()  # Lê o conteúdo como bytes
        return content.decode('ISO-8859-1', errors='replace')  # Usa 'ISO-8859-1' ou 'latin1' com 'replace' para evitar erros


@bot.command(name="test", help="Comando de teste para verificar se o bot está funcionando.")
async def test(ctx):
    await ctx.send("O bot está funcionando corretamente!")
    
@bot.command(name="coletar_links", help="Coleta links de páginas específicas. Uso: !coletar_links <url> <categorias>")
async def coletar_links(ctx, url: str, *categorias: str):
    if not categorias:
        categorias = ["/category/servicos-ao-cidadao", "category/noticias-da-secretaria", "category/modulo-destaques-da-secretaria", "category/noticias/todas-as-noticias", "noticias", "category/noticias", "category/modulo-destaques-da-secretaria", "category/modulo-carrossel-de-destaques-principais", "category/noticias-sem-fotos"]

    all_links = []
    loading_message = await ctx.send("Coletando links...")
    start_time = time.time()
    links_por_categoria = {}

    for categoria in categorias:
        categoria_url = f"{url.rstrip('/')}/{categoria}"
        page = 1
        links_categoria = []

        while True:
            if "saude.df.gov.br" in categoria_url:
                complete_url = f"{categoria_url}?p_p_id=com_liferay_asset_publisher_web_portlet_AssetPublisherPortlet_INSTANCE_Cziz3oWq1x3L&p_p_lifecycle=0&p_p_state=normal&p_p_mode=view&_com_liferay_asset_publisher_web_portlet_AssetPublisherPortlet_INSTANCE_Cziz3oWq1x3L_delta=10&p_r_p_resetCur=false&_com_liferay_asset_publisher_web_portlet_AssetPublisherPortlet_INSTANCE_Cziz3oWq1x3L_cur={page}" if page > 1 else categoria_url
            else:
                complete_url = f"{categoria_url}?page={page}" if page > 1 else categoria_url

            logging.info(f"Acessando a página: {complete_url}")
            try:
                links, current_page, total_pages = await extrair_links(complete_url)
                links_categoria.extend(links)
                logging.info(f"Encontrados {len(links)} links na página {current_page} para a categoria {categoria}.")
                if not links:
                    break
                page += 1
            except Exception as e:
                logging.error(f"Erro ao processar a página {page} na categoria {categoria}: {e}")
                break

        links_por_categoria[categoria] = links_categoria
        all_links.extend(links_categoria)

    elapsed_time = time.time() - start_time

    if all_links:
        output = io.BytesIO()
        data = [(categoria, link) for categoria, links in links_por_categoria.items() for link in links]
        df = pd.DataFrame(data, columns=["Categoria", "Links"])
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)

        mensagem_detalhada = "### Links coletados por categoria ###\n"
        for categoria, links in links_por_categoria.items():
            mensagem_detalhada += f"- **{categoria}**: {len(links)} links\n"
        mensagem_detalhada += f"\n**Total de links coletados**: {len(all_links)}\n"
        mensagem_detalhada += f"**Tempo gasto**: {elapsed_time:.2f} segundos."

        await loading_message.edit(content=mensagem_detalhada)
        await ctx.send(file=discord.File(output, filename="links_coletados.xlsx"))
    else:
        await loading_message.edit(content="Nenhum link encontrado para as categorias especificadas.")

async def extrair_links(url):
    try:
        response = await get_page(url)
        soup = BeautifulSoup(response, 'html.parser')

        links = [link['href'] for link in soup.select('.link-cor-de-coco') if link.get('href')]
        pagination = soup.select('.pagination a.page-numbers')
        current_page = 1
        total_pages = 1

        if pagination:
            for elem in pagination:
                if 'current' in (elem.get('class') or []):
                    current_page = int(elem.text.strip())
                    break
            total_pages = int(pagination[-1].text.strip().replace('.', ''))

        if links:
            return links, current_page, total_pages

        noticia_links = [link['href'] for link in soup.select('h4.entry-title a') if link.get('href')]
        pagination = soup.select('.pagination .page-item')
        current_page = 1
        total_pages = 1

        if pagination:
            for elem in pagination:
                link = elem.find('a')
                if link and 'cur=' in link.get('href', ''):
                    try:
                        current_page = int(link.get('href').split('cur=')[-1])
                        break
                    except ValueError:
                        current_page = 1
                        break

            try:
                total_pages = max(
                    int(link.get('href').split('cur=')[-1].replace('.', ''))
                    for elem in pagination if 'cur=' in (link := elem.find('a')).get('href', '')
                )
            except ValueError:
                total_pages = 1

        return noticia_links, current_page, total_pages

    except aiohttp.ClientError as e:
        logging.error(f"Erro ao extrair links da URL {url}: {e}")
        return [], 1, 1
    except Exception as e:
        logging.error(f"Erro inesperado ao extrair links da URL {url}: {e}")
        return [], 1, 1


    except aiohttp.ClientError as e:
        logging.error(f"Erro ao extrair links da URL {url}: {e}")
        return [], 1, 1
    except Exception as e:
        logging.error(f"Erro inesperado ao extrair links da URL {url}: {e}")
        return [], 1, 1

@bot.command(
    name="contar_arquivos", 
    help="Coleta arquivos de uma página específica. Uso: !contar_arquivos <url> [extrair_dominio (True/False)] [tipos_arquivos (ex: pdf,png)]"
)
@rate_limited
async def contar_arquivos(ctx, url: str, extrair_dominio: str = "False", tipos_arquivos: str = None):
    global client_session

    locais_especificos = ['portlet-layout', 'paginas-internas']
    domain_count = 0
    domain_name = extract_domain_name(url)
    url_count = {}

    logging.info(f"Dominio extraído: {domain_name}")

    extrair_dominio = extrair_dominio.lower() == 'true'
    additional_domains = ['seplad.df.gov.br', '10.242.0.138']

    tipos_arquivos = tipos_arquivos.split(',') if tipos_arquivos else ['pdf', 'png', 'jpeg', 'jpg', 'html', 'txt', 'docx', 'xlsx', 'mp3', 'mp4', 'zip', 'rar', 'pptx']

    loading_message = await ctx.send("Carregando arquivos...")
    start_time = time.time()

    def count_portal_links(soup, domain_name):
        """
        Conta links para páginas do portal, excluindo documentos e menu lateral.
        Retorna total de links do portal, links externos e lista de links encontrados.
        """
        total_links = 0
        internal_links = 0
        external_links = 0
        document_patterns = ['/documents/', '/wp-conteudo/', '/wp-content/']
        portal_links = []
        
        for local in locais_especificos:
            if local.startswith('.'):
                divs = soup.select(local)
            else:
                divs = soup.find_all('div', class_=local)

            for div in divs:
                # Ignorar menu lateral
                menu_lateral = div.find_all(class_='menu-lateral-flutuante')
                for menu in menu_lateral:
                    menu.decompose()

                # Processar links
                links = div.find_all('a')
                logging.info(f"Total de links encontrados na div: {len(links)}")

                for link in links:
                    href = link.get('href', '')
                    if not href:
                        logging.debug("Link ignorado: href vazio ou inexistente.")
                        continue

                    # Ignora links que começam com padrões de documento
                    if any(pattern in href for pattern in document_patterns):
                        logging.debug(f"Link ignorado (documento): {href}")
                        continue

                    # Verifica links internos e outros domínios permitidos
                    if href.startswith('/') and not any(pattern in href for pattern in document_patterns):
                        total_links += 1
                        internal_links += 1
                        portal_links.append(href)
                        logging.info(f"Link interno encontrado (relativo): {href}")
                    elif any(domain in href for domain in [domain_name] + additional_domains) and not any(pattern in href for pattern in document_patterns):
                        total_links += 1
                        internal_links += 1
                        portal_links.append(href)
                        logging.info(f"Link interno encontrado (absoluto, outro domínio): {href}")
                    elif href.startswith(('http://', 'https://')) and domain_name not in href and not any(domain in href for domain in additional_domains):
                        external_links += 1
                        logging.info(f"Link externo encontrado: {href}")

        logging.info(f"Resumo: Total links: {total_links}, Internos: {internal_links}, Externos: {external_links}")
        return total_links, internal_links, external_links, portal_links
    
    async def process_files(soup, url, domain_name):
        """
        Processa e coleta arquivos das páginas.
        Retorna lista de registros de arquivos e contagem do domínio.
        """
        records = []
        domain_count = 0
        valid_extensions = tuple(f'.{ext}' for ext in tipos_arquivos)
        tasks = []

        for local in locais_especificos:
            if local.startswith('.'):
                divs = soup.select(local)
            else:
                divs = soup.find_all('div', class_=local)

            if not divs:
                logging.warning(f"Nenhum elemento correspondente a '{local}' foi encontrado.")
                continue

            for div in divs:
                links = div.find_all('a')
                for link in links:
                    href = link.get('href', None)
                    if not href:
                        logging.warning("Link sem atributo 'href' encontrado, ignorando...")
                        continue

                    if href.startswith(('/documents', '/wp-content' '/wp-conteudo')):
                        href = urljoin(url, href)

                    file_url = urljoin(url, href) if not href.startswith(('http://', 'https://')) else href

                    base_url = file_url.split('?')[0]
                    if any(ext in base_url.lower() for ext in valid_extensions):
                        logging.info(f"Arquivo válido encontrado: {file_url}")
                        url_count[file_url] = url_count.get(file_url, 0) + 1
                        tasks.append(asyncio.create_task(check_and_record_url(file_url, records)))

                        extracted_domain = extract_domain_name(file_url).lower()
                        if extracted_domain.endswith(domain_name.lower()) or any(
                            extracted_domain.endswith(additional_domain.lower()) for additional_domain in additional_domains
                        ):
                            logging.info(f"Arquivo do domínio permitido encontrado: {file_url}")
                            domain_count += 1

        if tasks:
            await asyncio.gather(*tasks)
        return records, domain_count

    async def check_and_record_url(url, records, is_file=True):
        status_code = await check_url_status(url)
        is_duplicate = url_count.get(url, 0) > 1 if is_file else False
        records.append({
            'URL': url, 
            'Status': status_code,
            'Ocorrencias': url_count.get(url, 1),
            'Duplicada': 'Sim' if is_duplicate else 'Não',
            'Tipo': 'Arquivo' if is_file else 'Link'
        })

    async def check_url_status(url):
        try:
            async with client_session.head(url, allow_redirects=True, timeout=10, ssl=False) as response:
                return response.status
        except Exception as e:
            logging.error(f"Erro ao verificar o status da URL {url}: {e}")
            return 'Erro'

    try:
        if not url.startswith(('http://', 'https://')):
            await loading_message.edit(content="Erro: URL inválida. Certifique-se de incluir 'http://' ou 'https://'.")
            return

        html_content = await get_page(url)
        if html_content is None:
            await loading_message.edit(content="Erro ao obter conteúdo da página.")
            return

        # Processa o HTML uma única vez
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Conta links do portal
        total_links, internal_links, external_links, portal_links = count_portal_links(soup, domain_name)
        
        # Processa arquivos
        records, domain_count = await process_files(soup, url, domain_name)
        
        for link in portal_links:
            link_url = urljoin(url, link)
            await check_and_record_url(link_url, records, is_file=False)
        
        if not records:
            await loading_message.edit(content="Nenhum arquivo encontrado na página.")
            return

        records.sort(key=lambda x: (x['URL'], -x['Ocorrencias']))

        unique_records = []
        seen_urls = set()
        for record in records:
            if record['URL'] not in seen_urls:
                unique_records.append(record)
                seen_urls.add(record['URL'])

        df = pd.DataFrame(unique_records)
        expected_columns = ['URL', 'Status', 'Ocorrencias', 'Duplicada']

        actual_columns = df.columns.tolist()
        missing_columns = [col for col in expected_columns if col not in actual_columns]
        if missing_columns:
            logging.error(f"Colunas ausentes no DataFrame: {missing_columns}")
            await loading_message.edit(content=f"Erro: As seguintes colunas estão ausentes: {', '.join(missing_columns)}")
            return

        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']

            from openpyxl.styles import PatternFill
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            
            for row in range(2, len(df) + 2):
                tipo = worksheet.cell(row=row, column=5).value  # coluna 'Tipo'
                duplicada = worksheet.cell(row=row, column=4).value  # coluna 'Duplicada'
                
                if tipo == 'Link':
                    for col in range(1, 6):
                        worksheet.cell(row=row, column=col).fill = blue_fill
                elif duplicada == 'Sim':
                    for col in range(1, 6):
                        worksheet.cell(row=row, column=col).fill = yellow_fill

        output.seek(0)

        elapsed_time = time.time() - start_time
        num_collected = len(records)
        num_unique = len(unique_records)
        num_duplicates = num_collected - num_unique

        await loading_message.edit(content=f"Coleta finalizada! \n\n"
                                         f"Total de links para páginas do portal: **{total_links}**\n"
                                         f"Links internos do portal: **{internal_links}**\n"
                                         f"Links externos: **{external_links}**\n\n"
                                         f"Total de arquivos coletados: **{num_collected}**\n"
                                         f"Arquivos únicos: **{num_unique}**\n"
                                         f"Arquivos duplicados: **{num_duplicates}**\n"
                                         f"Total de arquivos do domínio **{domain_name}**: **{domain_count}**\n"
                                         f"Tempo gasto: **{elapsed_time:.2f} segundos**\n\n"
                                         f"Aguardando o envio do arquivo...")
        
        await ctx.send(file=discord.File(output, filename="arquivos_coletados.xlsx"))

    except Exception as e:
        error_message = f"Erro ao processar a página: {str(e)}"
        logging.error(error_message)
        await loading_message.edit(content=error_message)

def extract_domain_name(url):
    parsed_url = urlparse(url)
    domain = parsed_url.netloc.lower()
    if domain.startswith("www."):
        domain = domain[4:]
    return domain

def clean_url(url, valid_extensions):
    for ext in valid_extensions:
        if url.lower().endswith(ext):
            return url
    return None

# Inicia o bot
bot.run(DISCORD_TOKEN)